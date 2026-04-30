#!/usr/bin/env python3
"""Live E2E test for moj_isa_faq_to_excel.py.

This intentionally hits the official ISA/MOJ pages and verifies that the output
workbook is readable and contains non-empty Q&A rows for all eight FAQ pages.
"""

from __future__ import annotations

import subprocess
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SCRIPT = ROOT / "scripts" / "moj_isa_faq_to_excel.py"
MIN_EXPECTED_ROWS = 400
EXPECTED_PAGES = 8


def main() -> int:
    with tempfile.TemporaryDirectory(prefix="moj-isa-faq-e2e-") as tmpdir:
        output = Path(tmpdir) / "moj_isa_faq.xlsx"
        cmd = [
            sys.executable,
            str(SCRIPT),
            "--output",
            str(output),
            "--sleep",
            "0",
            "--expected-page-count",
            str(EXPECTED_PAGES),
            "--min-total-qa",
            str(MIN_EXPECTED_ROWS),
        ]
        completed = subprocess.run(cmd, cwd=ROOT, text=True, capture_output=True)
        print(completed.stdout, end="")
        if completed.stderr:
            print(completed.stderr, file=sys.stderr, end="")
        if completed.returncode != 0:
            return completed.returncode
        if not output.exists() or output.stat().st_size == 0:
            raise AssertionError("Excel output was not created")

        workbook = load_workbook(output, read_only=True, data_only=True)
        for sheet in ["QA", "Pages", "Summary"]:
            if sheet not in workbook.sheetnames:
                raise AssertionError(f"Missing sheet: {sheet}")

        qa_sheet = workbook["QA"]
        headers = [cell.value for cell in next(qa_sheet.iter_rows(min_row=1, max_row=1))]
        required_headers = {
            "page_no",
            "faq_page_title",
            "category",
            "section",
            "question_no",
            "question",
            "answer",
            "faq_page_url",
            "answer_page_url",
        }
        missing = required_headers.difference(headers)
        if missing:
            raise AssertionError(f"Missing QA columns: {sorted(missing)}")

        rows = list(qa_sheet.iter_rows(min_row=2, values_only=True))
        if len(rows) < MIN_EXPECTED_ROWS:
            raise AssertionError(f"Expected at least {MIN_EXPECTED_ROWS} QA rows, got {len(rows)}")

        idx = {name: headers.index(name) for name in required_headers}
        page_titles = {row[idx["faq_page_title"]] for row in rows if row[idx["faq_page_title"]]}
        if len(page_titles) != EXPECTED_PAGES:
            raise AssertionError(f"Expected {EXPECTED_PAGES} FAQ page titles, got {len(page_titles)}")

        empty_rows = [
            pos
            for pos, row in enumerate(rows, start=2)
            if not row[idx["question_no"]] or not row[idx["question"]] or not row[idx["answer"]]
        ]
        if empty_rows:
            raise AssertionError(f"Found empty Q/A cells at rows: {empty_rows[:10]}")

        print(f"E2E OK: {len(rows)} rows, {len(page_titles)} pages, workbook={output}")
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
