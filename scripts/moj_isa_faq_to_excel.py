#!/usr/bin/env python3
"""Scrape ISA/MOJ FAQ pages under qa_index.html and export Q&A rows to Excel.

Dependencies intentionally stay small: beautifulsoup4, pandas, openpyxl.
Network access uses urllib from the Python standard library.
"""

from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
import re
import sys
import time
from collections import OrderedDict, defaultdict
from pathlib import Path
from typing import Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urldefrag, urljoin, urlparse
from urllib.request import Request, urlopen

import pandas as pd
from bs4 import BeautifulSoup, Tag
from openpyxl.styles import Alignment, Font, PatternFill

INDEX_URL = "https://www.moj.go.jp/isa/applications/faq/qa_index.html"
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0 Safari/537.36"
)
REQUEST_HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
}
FULLWIDTH_DIGITS = str.maketrans("０１２３４５６７８９", "0123456789")
QUESTION_RE = re.compile(
    r"^[\s\u200b\ufeff　]*(?:Q|Ｑ)\s*"
    r"([0-9０-９]+(?:[\-－−ー][0-9０-９]+)?)"
    r"\s*[：:．\.]?\s*(.*)$"
)
ANSWER_RE = re.compile(r"^[\s\u200b\ufeff　]*(?:A|Ａ)\s*[：:．\.]?\s*(.*)$")
ANSWER_FRAGMENT_RE = re.compile(r"^q[0-9０-９]+(?:[\-－−ー][0-9０-９]+)?-a$", re.I)
HEADING_TAGS = ["h1", "h2", "h3", "h4", "h5", "h6"]
CONTENT_SELECTORS = ["#contentsArea", "#tex", "main", "body"]


class ScrapeError(RuntimeError):
    """Raised when scraping cannot produce trustworthy output."""


@dataclasses.dataclass(frozen=True)
class FaqPage:
    order: int
    title: str
    url: str


@dataclasses.dataclass
class QaRow:
    page_no: int
    faq_page_title: str
    category: str
    section: str
    question_no: str
    question: str
    answer: str
    faq_page_url: str
    answer_page_url: str

    def as_dict(self) -> dict[str, object]:
        return dataclasses.asdict(self)


def fetch_html(url: str, timeout: int) -> str:
    request = Request(url, headers=REQUEST_HEADERS)
    try:
        with urlopen(request, timeout=timeout) as response:
            raw = response.read()
    except HTTPError as exc:
        raise ScrapeError(f"HTTP error while fetching {url}: {exc.code} {exc.reason}") from exc
    except URLError as exc:
        raise ScrapeError(f"Network error while fetching {url}: {exc.reason}") from exc
    except TimeoutError as exc:
        raise ScrapeError(f"Timeout while fetching {url}") from exc

    # The ISA site declares UTF-8. Keep replacement explicit so a single odd byte
    # does not hide the rest of the page.
    return raw.decode("utf-8", errors="replace")


def soup_from_html(html: str) -> BeautifulSoup:
    return BeautifulSoup(html, "html.parser")


def content_root(soup: BeautifulSoup | Tag) -> Tag:
    for selector in CONTENT_SELECTORS:
        found = soup.select_one(selector)
        if found is not None:
            return found
    if isinstance(soup, Tag):
        return soup
    raise ScrapeError("No content root found in HTML")


def normalize_space(text: str) -> str:
    text = text.replace("\ufeff", "").replace("\u200b", "")
    text = text.replace("\xa0", " ").replace("　", " ")
    return re.sub(r"[ \t\r\f\v]+", " ", text).strip()


def inline_text(node: Tag | str | None) -> str:
    if node is None:
        return ""
    if isinstance(node, str):
        return normalize_space(node)
    return normalize_space(node.get_text(" ", strip=True))


def block_text(node: Tag | str | None) -> str:
    if node is None:
        return ""
    if isinstance(node, str):
        raw = node
    else:
        raw = node.get_text("\n", strip=True)
    lines = [normalize_space(line) for line in raw.splitlines()]
    lines = [line for line in lines if line and line not in {"▲ ページトップへ", "▲ 質問リストへ"}]
    return "\n".join(lines)


def normalize_q_no(raw: str) -> str:
    value = raw.translate(FULLWIDTH_DIGITS)
    value = value.replace("－", "-").replace("−", "-").replace("ー", "-")
    return f"Q{value}"


def parse_question(text: str) -> tuple[str, str] | None:
    m = QUESTION_RE.match(inline_text(text))
    if not m:
        return None
    return normalize_q_no(m.group(1)), normalize_space(m.group(2))


def parse_answer(text: str) -> str | None:
    m = ANSWER_RE.match(inline_text(text))
    if not m:
        return None
    return normalize_space(m.group(1))


def strip_q_marker(text: str) -> str:
    parsed = parse_question(text)
    return parsed[1] if parsed else inline_text(text)


def strip_answer_marker(text: str) -> str:
    parsed = parse_answer(text)
    return parsed if parsed is not None else inline_text(text)


def document_title(soup: BeautifulSoup | Tag) -> str:
    root = content_root(soup)
    h1 = root.find("h1")
    if h1:
        return inline_text(h1)
    title = soup.find("title") if isinstance(soup, BeautifulSoup) else None
    return inline_text(title) if title else ""


def normalize_heading(text: str) -> str:
    text = normalize_space(text)
    text = re.sub(r"[─\-–—]+\s*(Question|Answer)\s*[─\-–—]+", "", text, flags=re.I)
    text = text.replace("答えパート開始", "").strip()
    return normalize_space(text)


def previous_heading_text(node: Tag) -> str:
    heading = node.find_previous(HEADING_TAGS)
    if heading is None:
        return ""
    text = normalize_heading(inline_text(heading))
    if text in {"答え", "Q＆A", "Ｑ＆Ａ"}:
        return ""
    return text


def heading_path(levels: dict[int, str]) -> str:
    return " > ".join(value for _, value in sorted(levels.items()) if value)


def build_question_section_index(soup: BeautifulSoup | Tag) -> dict[str, str]:
    """Build Q number -> section mapping from question lists.

    Several pages put a question list first and answers later or on another page.
    Keeping this map lets the Excel preserve the original category even when the
    answer block itself only says "Answer".
    """
    root = content_root(soup)
    mapping: dict[str, str] = {}
    levels: dict[int, str] = {}
    pseudo_subsection = ""

    for node in root.find_all(HEADING_TAGS + ["p", "dl"], recursive=True):
        if node.name in HEADING_TAGS:
            level = int(node.name[1])
            text = normalize_heading(inline_text(node))
            if not text or "Answer" in text or text == "答え":
                continue
            levels = {k: v for k, v in levels.items() if k < level}
            levels[level] = text
            pseudo_subsection = ""
            continue

        text = inline_text(node)
        if re.fullmatch(r"＜[^＞]+＞", text):
            pseudo_subsection = text
            continue

        q_no = None
        if node.name == "dl":
            dt_node = node.find("dt")
            dt_text = inline_text(dt_node)
            parsed = parse_question(dt_text)
            if parsed:
                q_no = parsed[0]
        else:
            parsed = parse_question(text)
            if parsed:
                q_no = parsed[0]

        if q_no:
            section = heading_path(levels)
            if pseudo_subsection:
                section = f"{section} > {pseudo_subsection}" if section else pseudo_subsection
            mapping.setdefault(q_no, section)

    return mapping


def discover_faq_pages(index_url: str, timeout: int) -> list[FaqPage]:
    soup = soup_from_html(fetch_html(index_url, timeout))
    root = content_root(soup)
    pages: list[FaqPage] = []
    seen: set[str] = set()

    for a in root.select("ul.menuList01 a[href]"):
        title = inline_text(a)
        href = urljoin(index_url, a.get("href", ""))
        href, _fragment = urldefrag(href)
        if not title or href in seen:
            continue
        seen.add(href)
        pages.append(FaqPage(order=len(pages) + 1, title=title, url=href))

    return pages


def same_site(url_a: str, url_b: str) -> bool:
    a = urlparse(url_a)
    b = urlparse(url_b)
    return a.scheme == b.scheme and a.netloc == b.netloc


def discover_answer_document_urls(page_url: str, soup: BeautifulSoup | Tag) -> list[str]:
    root = content_root(soup)
    urls: OrderedDict[str, None] = OrderedDict()
    page_base, _ = urldefrag(page_url)

    for a in root.select("a[href]"):
        raw_href = a.get("href") or ""
        href = urljoin(page_url, raw_href)
        href_base, fragment = urldefrag(href)
        if not same_site(page_url, href_base):
            continue
        if href_base == page_base:
            continue
        if ANSWER_FRAGMENT_RE.match(fragment) or "q-and-a_page" in href_base:
            urls.setdefault(href_base, None)

    return list(urls.keys())


def first_direct_or_nested(parent: Tag, name: str) -> Tag | None:
    found = parent.find(name, recursive=False)
    return found if found is not None else parent.find(name)


def answer_fragment_url(base_url: str, node: Tag, fallback_id: str = "") -> str:
    for tag in [node, *node.find_all(True)]:
        tag_id = tag.get("id")
        if tag_id:
            return f"{base_url}#{tag_id}"
    if fallback_id:
        return f"{base_url}#{fallback_id}"
    return base_url


def row_from_parts(
    *,
    page: FaqPage,
    category: str,
    section: str,
    q_no: str,
    question: str,
    answer: str,
    answer_url: str,
) -> QaRow | None:
    question = normalize_space(question)
    answer = block_text(answer)
    if not q_no or not question or not answer:
        return None
    return QaRow(
        page_no=page.order,
        faq_page_title=page.title,
        category=category,
        section=section,
        question_no=q_no,
        question=question,
        answer=answer,
        faq_page_url=page.url,
        answer_page_url=answer_url,
    )


def extract_qa_boxes(page: FaqPage, soup: BeautifulSoup | Tag, doc_url: str, q_section: dict[str, str]) -> list[QaRow]:
    root = content_root(soup)
    rows: list[QaRow] = []
    pending_question: dict[str, str] | None = None

    for box in root.select(".qa_box"):
        a_box = box.select_one(".a_box")

        q_dl = None
        for dl in box.find_all("dl"):
            if dl.find_parent(class_="a_box") is None:
                dt_node = dl.find("dt")
                if dt_node is not None and parse_question(inline_text(dt_node)):
                    q_dl = dl
                    break

        q_info: dict[str, str] | None = None
        if q_dl is not None:
            q_dt = q_dl.find("dt")
            q_dd = q_dl.find("dd")
            parsed = parse_question(inline_text(q_dt))
            if parsed:
                q_no, q_text = parsed
                if not q_text:
                    q_text = inline_text(q_dd)
                if q_text:
                    q_info = {
                        "q_no": q_no,
                        "question": q_text,
                        "section": q_section.get(q_no) or previous_heading_text(box),
                        "answer_url": answer_fragment_url(doc_url, box, q_no.lower()),
                    }

        answer = ""
        if a_box is not None:
            a_dd = None
            for answer_dl in a_box.find_all("dl"):
                a_dd = answer_dl.find("dd")
                if a_dd is not None:
                    break
            if a_dd is None:
                a_dd = a_box.find("dd")
            answer = block_text(a_dd)

        if q_info and answer:
            row = row_from_parts(
                page=page,
                category="",
                section=q_info["section"],
                q_no=q_info["q_no"],
                question=q_info["question"],
                answer=answer,
                answer_url=q_info["answer_url"],
            )
            if row:
                rows.append(row)
            pending_question = None
        elif q_info and not answer:
            # Some legacy pages split one Q&A over two consecutive .qa_box
            # elements: the first has only the question, the second only A.
            pending_question = q_info
        elif pending_question and answer:
            row = row_from_parts(
                page=page,
                category="",
                section=pending_question["section"],
                q_no=pending_question["q_no"],
                question=pending_question["question"],
                answer=answer,
                answer_url=pending_question["answer_url"],
            )
            if row:
                rows.append(row)
            pending_question = None

    return rows


def extract_dl_qas(page: FaqPage, soup: BeautifulSoup | Tag, doc_url: str, q_section: dict[str, str]) -> list[QaRow]:
    root = content_root(soup)
    rows: list[QaRow] = []
    current: dict[str, str | Tag] | None = None

    for node in root.find_all(["dt", "dd"], recursive=True):
        if node.find_parent(class_="qa_box") is not None:
            continue

        if node.name == "dt":
            parsed = parse_question(inline_text(node))
            if not parsed:
                continue
            q_no, q_text = parsed
            # Question-list pages often use <dt>Q1:</dt><dd><a>question</a></dd>.
            # A real answer block has the question text in dt, even when the
            # following <dd> lives in another malformed <dl>.
            if not q_text:
                current = None
                continue
            current = {
                "q_no": q_no,
                "question": q_text,
                "section": q_section.get(q_no) or previous_heading_text(node),
                "anchor_node": node.find_parent("dl") or node,
            }
            continue

        if node.name == "dd" and current is not None:
            row = row_from_parts(
                page=page,
                category="",
                section=str(current.get("section", "")),
                q_no=str(current["q_no"]),
                question=str(current["question"]),
                answer=block_text(node),
                answer_url=answer_fragment_url(doc_url, current["anchor_node"], str(current["q_no"]).lower()),
            )
            if row:
                rows.append(row)
            current = None

    return rows


def tab_labels(root: Tag) -> list[str]:
    labels = []
    for item in root.select(".tab2 .tab2__item"):
        labels.append(inline_text(item))
    return labels


def iter_paragraph_blocks(container: Tag) -> Iterable[Tag]:
    for node in container.find_all(["h2", "h3", "h4", "h5", "p"], recursive=True):
        if node.find_parent(class_="bannerBox") is not None:
            continue
        yield node


def extract_tabbed_paragraph_qas(page: FaqPage, soup: BeautifulSoup | Tag, doc_url: str) -> list[QaRow]:
    root = content_root(soup)
    bodies = root.select(".tab2-body__item")
    if not bodies:
        return []

    labels = tab_labels(root)
    rows: list[QaRow] = []

    for idx, body in enumerate(bodies):
        category = labels[idx] if idx < len(labels) else f"タブ{idx + 1}"
        section = ""
        current: dict[str, str] | None = None
        answer_parts: list[str] = []
        anchor_url = doc_url

        def flush() -> None:
            nonlocal current, answer_parts, anchor_url
            if current is None:
                return
            row = row_from_parts(
                page=page,
                category=category,
                section=current.get("section", ""),
                q_no=current["q_no"],
                question=current["question"],
                answer="\n".join(answer_parts),
                answer_url=anchor_url,
            )
            if row:
                rows.append(row)
            current = None
            answer_parts = []
            anchor_url = doc_url

        for node in iter_paragraph_blocks(body):
            if node.name in HEADING_TAGS:
                text = normalize_heading(inline_text(node))
                if text:
                    section = text
                continue

            text = inline_text(node)
            parsed_q = parse_question(text)
            if parsed_q:
                flush()
                q_no, question = parsed_q
                current = {"q_no": q_no, "question": question, "section": section}
                anchor_url = answer_fragment_url(doc_url, node, q_no.lower())
                continue

            if current is None:
                continue
            answer_piece = strip_answer_marker(text)
            if answer_piece:
                answer_parts.append(answer_piece)

        flush()

    return rows


def extract_id_paragraph_qas(page: FaqPage, soup: BeautifulSoup | Tag, doc_url: str, q_section: dict[str, str]) -> list[QaRow]:
    root = content_root(soup)
    q_nodes = [p for p in root.find_all("p") if p.get("id") and parse_question(inline_text(p))]
    if not q_nodes:
        return []

    q_node_ids = {id(node) for node in q_nodes}
    rows: list[QaRow] = []
    current: dict[str, str] | None = None
    answer_parts: list[str] = []
    anchor_url = doc_url
    section = ""

    def flush() -> None:
        nonlocal current, answer_parts, anchor_url
        if current is None:
            return
        row = row_from_parts(
            page=page,
            category="",
            section=current.get("section", ""),
            q_no=current["q_no"],
            question=current["question"],
            answer="\n".join(answer_parts),
            answer_url=anchor_url,
        )
        if row:
            rows.append(row)
        current = None
        answer_parts = []
        anchor_url = doc_url

    for node in iter_paragraph_blocks(root):
        if node.name in HEADING_TAGS:
            text = normalize_heading(inline_text(node))
            if text and text != "答え":
                section = text
            continue

        text = inline_text(node)
        if id(node) in q_node_ids:
            flush()
            parsed = parse_question(text)
            if not parsed:
                continue
            q_no, question = parsed
            current = {"q_no": q_no, "question": question, "section": q_section.get(q_no) or section}
            anchor_url = answer_fragment_url(doc_url, node, q_no.lower())
            continue

        if current is None:
            continue
        answer_piece = strip_answer_marker(text)
        if answer_piece:
            answer_parts.append(answer_piece)

    flush()
    return rows


def extract_rows_from_document(page: FaqPage, soup: BeautifulSoup | Tag, doc_url: str, q_section: dict[str, str]) -> list[QaRow]:
    rows: list[QaRow] = []
    rows.extend(extract_qa_boxes(page, soup, doc_url, q_section))
    rows.extend(extract_dl_qas(page, soup, doc_url, q_section))
    rows.extend(extract_tabbed_paragraph_qas(page, soup, doc_url))
    rows.extend(extract_id_paragraph_qas(page, soup, doc_url, q_section))
    return rows


def scrape(index_url: str, timeout: int, expected_page_count: int, sleep_seconds: float) -> tuple[list[QaRow], list[dict[str, object]]]:
    pages = discover_faq_pages(index_url, timeout)
    if len(pages) != expected_page_count:
        raise ScrapeError(f"Expected {expected_page_count} FAQ pages under index, found {len(pages)}")

    all_rows: list[QaRow] = []
    page_summaries: list[dict[str, object]] = []
    seen: set[tuple[str, str, str, str, str]] = set()

    for page in pages:
        html = fetch_html(page.url, timeout)
        soup = soup_from_html(html)
        q_section = build_question_section_index(soup)
        answer_doc_urls = [page.url] + discover_answer_document_urls(page.url, soup)
        doc_counts: dict[str, int] = defaultdict(int)
        page_count_before = len(all_rows)

        for doc_index, doc_url in enumerate(answer_doc_urls):
            doc_soup = soup if doc_index == 0 else soup_from_html(fetch_html(doc_url, timeout))
            for row in extract_rows_from_document(page, doc_soup, doc_url, q_section):
                key = (row.faq_page_url, row.answer_page_url, row.category, row.question_no, row.question)
                if key in seen:
                    continue
                seen.add(key)
                all_rows.append(row)
                doc_counts[doc_url] += 1
            if sleep_seconds > 0 and doc_index < len(answer_doc_urls) - 1:
                time.sleep(sleep_seconds)

        page_row_count = len(all_rows) - page_count_before
        if page_row_count <= 0:
            raise ScrapeError(f"No Q&A rows extracted from {page.title}: {page.url}")

        page_summaries.append(
            {
                "page_no": page.order,
                "faq_page_title": page.title,
                "faq_page_url": page.url,
                "answer_documents": len(answer_doc_urls),
                "qa_rows": page_row_count,
                "answer_document_urls": "\n".join(answer_doc_urls),
            }
        )
        if sleep_seconds > 0:
            time.sleep(sleep_seconds)

    return all_rows, page_summaries


def write_excel(rows: list[QaRow], page_summaries: list[dict[str, object]], output_path: Path, index_url: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    qa_df = pd.DataFrame([row.as_dict() for row in rows])
    pages_df = pd.DataFrame(page_summaries)
    summary_df = pd.DataFrame(
        [
            {"key": "generated_at", "value": dt.datetime.now().isoformat(timespec="seconds")},
            {"key": "index_url", "value": index_url},
            {"key": "faq_pages", "value": len(page_summaries)},
            {"key": "qa_rows", "value": len(rows)},
        ]
    )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        qa_df.to_excel(writer, sheet_name="QA", index=False)
        pages_df.to_excel(writer, sheet_name="Pages", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)

        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for sheet_name in ["QA", "Pages", "Summary"]:
            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

        qa_widths = {
            "A": 8,
            "B": 30,
            "C": 28,
            "D": 32,
            "E": 10,
            "F": 60,
            "G": 90,
            "H": 45,
            "I": 45,
        }
        for col, width in qa_widths.items():
            workbook["QA"].column_dimensions[col].width = width
        for ws_name in ["Pages", "Summary"]:
            ws = workbook[ws_name]
            for col in ws.columns:
                letter = col[0].column_letter
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 80)


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Scrape the eight ISA/MOJ FAQ pages under qa_index.html and export Q&A to Excel."
    )
    parser.add_argument("-o", "--output", required=True, help="Output .xlsx path")
    parser.add_argument("--index-url", default=INDEX_URL, help=f"FAQ index URL. Default: {INDEX_URL}")
    parser.add_argument("--timeout", type=int, default=30, help="HTTP timeout seconds. Default: 30")
    parser.add_argument("--sleep", type=float, default=0.2, help="Delay between page fetches. Default: 0.2")
    parser.add_argument("--expected-page-count", type=int, default=8, help="Fail unless this many FAQ pages are found. Default: 8")
    parser.add_argument("--min-total-qa", type=int, default=1, help="Fail unless at least this many Q&A rows are extracted. Default: 1")
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(sys.argv[1:] if argv is None else argv)
    try:
        rows, page_summaries = scrape(
            index_url=args.index_url,
            timeout=args.timeout,
            expected_page_count=args.expected_page_count,
            sleep_seconds=args.sleep,
        )
        if len(rows) < args.min_total_qa:
            raise ScrapeError(f"Expected at least {args.min_total_qa} Q&A rows, found {len(rows)}")
        output_path = Path(args.output)
        write_excel(rows, page_summaries, output_path, args.index_url)
    except ScrapeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Wrote {len(rows)} Q&A rows from {len(page_summaries)} FAQ pages to {output_path}")
    for summary in page_summaries:
        print(f"- {summary['faq_page_title']}: {summary['qa_rows']} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
